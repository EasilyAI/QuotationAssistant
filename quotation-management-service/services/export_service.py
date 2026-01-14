"""
Export service for generating Excel files.
"""

import os
import logging
from typing import Optional, Dict, Any
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from services.quotation_service import get_quotation, get_quotations_table

logger = logging.getLogger(__name__)
log_level = os.getenv('LOG_LEVEL', 'INFO').upper()
logger.setLevel(getattr(logging, log_level, logging.INFO))


def generate_stock_check_excel(quotation: Dict[str, Any]) -> BytesIO:
    """
    Generate stock check Excel file for manufacturer order list.
    
    Format: ordering_number, quantity (ONLY - no product name or other columns)
    
    Args:
        quotation: Quotation data
    
    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Check"
    
    # Header row - ONLY ordering number and quantity (no product name)
    headers = ['Ordering Number', 'Quantity']
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows - only lines with ordering_number
    # Include ONLY ordering_number and quantity (no product_name, description, etc.)
    lines = quotation.get('lines', [])
    for line in lines:
        ordering_number = line.get('ordering_number', '').strip()
        if ordering_number:  # Only include lines with ordering number
            quantity = line.get('quantity', 1)
            quantity_float = float(quantity) if quantity is not None else 1.0
            # Append ONLY ordering number and quantity - no other columns
            ws.append([
                ordering_number,
                quantity_float,
            ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_priority_import_excel(quotation: Dict[str, Any]) -> BytesIO:
    """
    Generate priority import Excel file for ERP ingestion.
    
    Format: ordering_number, quantity, final_price
    
    Args:
        quotation: Quotation data
    
    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Priority Import"
    
    # Header row - only ordering number, quantity, and price
    headers = [
        'Ordering Number',
        'Quantity',
        'Price'
    ]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows - only lines with ordering_number
    lines = quotation.get('lines', [])
    for line in lines:
        ordering_number = line.get('ordering_number', '').strip()
        if ordering_number:  # Only include lines with ordering number
            quantity = line.get('quantity', 1)
            final_price = line.get('final_price', 0.0)
            
            # Convert to float to handle Decimal types from DynamoDB
            quantity_float = float(quantity) if quantity is not None else 1.0
            final_price_float = float(final_price) if final_price is not None else 0.0
            
            ws.append([
                ordering_number,
                quantity_float,
                final_price_float
            ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


# Removed S3 upload functions - exports are returned directly for download


def export_stock_check(quotation_id: str) -> Optional[BytesIO]:
    """
    Generate stock check Excel export and return as BytesIO for direct download.
    
    Args:
        quotation_id: Quotation ID
    
    Returns:
        BytesIO object containing Excel file, or None on error
    """
    quotation = get_quotation(quotation_id)
    if not quotation:
        return None
    
    # Generate Excel
    excel_data = generate_stock_check_excel(quotation)
    
    # Update quotation with export timestamp (optional metadata)
    table = get_quotations_table()
    try:
        table.update_item(
            Key={'quotation_id': quotation_id},
            UpdateExpression="SET #exports.#last_exported_at = :timestamp, #updated_at = :updated_at",
            ExpressionAttributeNames={
                '#exports': 'exports',
                '#last_exported_at': 'last_exported_at',
                '#updated_at': 'updated_at'
            },
            ExpressionAttributeValues={
                ':timestamp': datetime.utcnow().isoformat() + "Z",
                ':updated_at': datetime.utcnow().isoformat() + "Z"
            }
        )
    except Exception as e:
        logger.error(f"Error updating quotation export info: {str(e)}")
    
    return excel_data


def export_priority_import(quotation_id: str) -> Optional[BytesIO]:
    """
    Generate priority import Excel export and return as BytesIO for direct download.
    
    Args:
        quotation_id: Quotation ID
    
    Returns:
        BytesIO object containing Excel file, or None on error
    """
    quotation = get_quotation(quotation_id)
    if not quotation:
        return None
    
    # Generate Excel
    excel_data = generate_priority_import_excel(quotation)
    
    # Update quotation with export timestamp (optional metadata)
    table = get_quotations_table()
    try:
        table.update_item(
            Key={'quotation_id': quotation_id},
            UpdateExpression="SET #exports.#last_exported_at = :timestamp, #updated_at = :updated_at",
            ExpressionAttributeNames={
                '#exports': 'exports',
                '#last_exported_at': 'last_exported_at',
                '#updated_at': 'updated_at'
            },
            ExpressionAttributeValues={
                ':timestamp': datetime.utcnow().isoformat() + "Z",
                ':updated_at': datetime.utcnow().isoformat() + "Z"
            }
        )
    except Exception as e:
        logger.error(f"Error updating quotation export info: {str(e)}")
    
    return excel_data


# Removed get_export_download_url - exports are generated on-demand and returned directly

