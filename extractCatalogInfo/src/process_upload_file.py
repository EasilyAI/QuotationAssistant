import json
import os
import uuid
import time
import re
import tempfile
from datetime import datetime

import boto3
from openpyxl import load_workbook

from utils.incomingEventParser import parse_s3_key
from utils.corsHeaders import get_cors_headers
from utils.helpers import convert_floats_to_decimal

s3 = boto3.client("s3")
dynamodb = boto3.resource("dynamodb")
s3_client = boto3.client("s3")

BUCKET = "hb-files-raw"
AWS_REGION = "us-east-1"
# BUCKET = os.environ["UPLOAD_BUCKET"]
FILES_TABLE = os.environ["FILES_TABLE"]
CATALOG_PRODUCTS_TABLE = os.environ.get("CATALOG_PRODUCTS_TABLE", "hb-catalog-products")
PRICE_LIST_PRODUCTS_TABLE = os.environ.get("PRICE_LIST_PRODUCTS_TABLE", "hb-price-list-products")

# Expected schema for price list files
PRICE_LIST_SCHEMA = {
    "columns": [
        {"name": "orderingNumber", "type": "string", "required": True},
        {"name": "description", "type": "string", "required": True},
        {"name": "price", "type": "number", "required": True},
        # Optional Swagelok product link column
        {"name": "SwagelokLink", "type": "link", "required": False},
    ]
}

def update_file_status(file_id, status, **kwargs):
    """
    Update file processing status in DynamoDB.
    
    Args:
        file_id: File ID to update
        status: Processing status (processing, completed, failed, etc.)
        **kwargs: Additional attributes to update (metadata, error messages, etc.)
    """
    table = dynamodb.Table(FILES_TABLE)
    
    update_expression_parts = ["#status = :status", "#updatedAt = :updatedAt"]
    expression_attribute_names = {
        "#status": "status",
        "#updatedAt": "updatedAt"
    }
    expression_attribute_values = {
        ":status": status,
        ":updatedAt": int(time.time() * 1000)  # Timestamp in milliseconds
    }
    
    # Add any additional attributes from kwargs
    for key, value in kwargs.items():
        attr_name = f"#{key}"
        attr_value = f":{key}"
        update_expression_parts.append(f"{attr_name} = {attr_value}")
        expression_attribute_names[attr_name] = key
        expression_attribute_values[attr_value] = value
    
    update_expression = "SET " + ", ".join(update_expression_parts)
    
    try:
        table.update_item(
            Key={"fileId": file_id},
            UpdateExpression=update_expression,
            ExpressionAttributeNames=expression_attribute_names,
            ExpressionAttributeValues=expression_attribute_values
        )
        print(f"[update_file_status] Updated file {file_id} with status: {status}")
    except Exception as e:
        print(f"[update_file_status] ERROR: Failed to update file status: {e}")


def save_products_to_catalog_products_table(file_id, s3_key, event_payloads):
    """
    Save extracted products to catalog products table for review.
    Stores ALL products as ONE document with fileId as the primary key.
    
    Args:
        file_id: File ID associated with these products
        s3_key: S3 key of the source file
        event_payloads: List of dictionaries containing products keyed by ordering number
    
    Returns:
        int: Number of products saved
    """
    table = dynamodb.Table(CATALOG_PRODUCTS_TABLE)
    timestamp = int(time.time() * 1000)
    
    print(f"[save_products_to_catalog_products_table] Starting to process {len(event_payloads)} event payloads")
    
    # Build a single products list from all event payloads
    all_products = []
    
    for table_idx, products_dict in enumerate(event_payloads):
        print(f"[save_products_to_catalog_products_table] Processing event_payload {table_idx}: {len(products_dict)} products in this payload")
        
        for ordering_number, product_data in products_dict.items():
            try:
                print(f"[save_products_to_catalog_products_table] Processing product: {ordering_number}")
                
                # Convert any float values to Decimal for DynamoDB
                product_item = {
                    "orderingNumber": ordering_number,
                    "tableIndex": table_idx,
                    "status": "pending_review",  # pending_review, approved, rejected
                }
                
                # Add specs if present (convert floats to Decimal)
                if "specs" in product_data and product_data["specs"]:
                    product_item["specs"] = convert_floats_to_decimal(product_data["specs"])
                    print(f"[save_products_to_catalog_products_table]   - Added {len(product_data['specs'])} specs")
                
                # Add location if present (convert floats to Decimal)
                if "location" in product_data and product_data["location"]:
                    product_item["location"] = convert_floats_to_decimal(product_data["location"])
                
                # Add tindex if present
                if "tindex" in product_data:
                    product_item["tindex"] = product_data["tindex"]
                
                # Add id if present
                if "id" in product_data:
                    product_item["id"] = product_data["id"]
                
                all_products.append(product_item)
                print(f"[save_products_to_catalog_products_table]   - Product added successfully. Total products so far: {len(all_products)}")
                
            except Exception as e:
                print(f"[save_products_to_catalog_products_table] ERROR: Failed to process product {ordering_number}: {e}")
                print(f"[save_products_to_catalog_products_table] Product data: {product_data}")
                import traceback
                traceback.print_exc()
    
    print(f"[save_products_to_catalog_products_table] Finished processing all payloads. Total products collected: {len(all_products)}")
    
    # Save all products as ONE document with fileId as the key
    try:
        document = {
            "fileId": file_id,
            "sourceFile": s3_key,
            "createdAt": timestamp,
            "updatedAt": timestamp,
            "products": all_products,
            "productsCount": len(all_products)
        }
        
        print(f"[save_products_to_catalog_products_table] Attempting to save document with {len(all_products)} products to DynamoDB...")
        table.put_item(Item=document)
        print(f"[save_products_to_catalog_products_table] SUCCESS: Saved {len(all_products)} products as one document with fileId: {file_id}")
        return len(all_products)
        
    except Exception as e:
        print(f"[save_products_to_catalog_products_table] ERROR: Failed to save products document: {e}")
        import traceback
        traceback.print_exc()
        return 0


def find_header_row(rows, max_rows_to_scan=20):
    """
    Search for the header row in the xlsx file by looking for expected column names.
    The header row should contain at least the required columns (orderingNumber, description, price).
    
    Args:
        rows: List of rows from the xlsx file
        max_rows_to_scan: Maximum number of rows to scan for headers
    
    Returns:
        tuple: (header_row_index: int or None, headers: list or None)
    """
    expected_col_names = [col["name"].lower() for col in PRICE_LIST_SCHEMA["columns"]]
    required_col_names = [col["name"].lower() for col in PRICE_LIST_SCHEMA["columns"] if col["required"]]
    
    for row_idx, row in enumerate(rows[:max_rows_to_scan]):
        if row is None:
            continue
            
        # Convert row to lowercase strings for comparison
        row_values = []
        for cell in row:
            if cell is not None:
                row_values.append(str(cell).strip().lower())
            else:
                row_values.append(None)
        
        if row_idx < 5:
            # Log first few rows to help debug header detection issues
            print(f"[find_header_row] Row {row_idx} values (lowercased): {row_values}")
        
        # Check if this row contains the required column names
        found_required = 0
        for req_col in required_col_names:
            if req_col in row_values:
                found_required += 1
        
        # If we found all required columns, this is likely the header row
        if found_required >= len(required_col_names):
            print(f"[find_header_row] Found header row at index {row_idx}: {row}")
            return row_idx, list(row)
    
    return None, None


def validate_price_list_schema(headers):
    """
    Validate that the price list file has the expected schema.
    
    Args:
        headers: List of header names from the xlsx file
    
    Returns:
        tuple: (is_valid: bool, errors: list of error messages, column_mapping: dict mapping expected names to actual indices)
    """
    expected_col_names = [col["name"] for col in PRICE_LIST_SCHEMA["columns"]]
    required_col_names = [col["name"] for col in PRICE_LIST_SCHEMA["columns"] if col["required"]]
    errors = []
    column_mapping = {}
    
    # Check if we have headers
    if not headers:
        return False, ["No headers found in the file"], {}
    
    # Normalize headers for comparison
    normalized_headers = []
    for h in headers:
        if h is not None:
            normalized_headers.append(str(h).strip().lower())
        else:
            normalized_headers.append(None)

    print(f"[validate_price_list_schema] Raw headers: {headers}")
    print(f"[validate_price_list_schema] Normalized headers: {normalized_headers}")
    
    # Find each expected column in the headers (flexible order)
    for expected_col in expected_col_names:
        expected_lower = expected_col.lower()
        found_idx = None
        
        for idx, header in enumerate(normalized_headers):
            if header == expected_lower:
                found_idx = idx
                break
        
        if found_idx is not None:
            column_mapping[expected_col] = found_idx
        elif expected_col in required_col_names:
            errors.append(f"Required column '{expected_col}' not found in headers")
    
    # Check we found all required columns
    for req_col in required_col_names:
        if req_col not in column_mapping:
            if f"Required column '{req_col}' not found in headers" not in errors:
                errors.append(f"Required column '{req_col}' not found in headers")
    
    is_valid = len(errors) == 0
    if not is_valid:
        print(f"[validate_price_list_schema] Schema INVALID. Errors: {errors}")
    else:
        print(f"[validate_price_list_schema] Schema valid. Column mapping will be based on: {column_mapping}")
    return is_valid, errors, column_mapping


def validate_price_list_row(row_data, row_number):
    """
    Validate a single row from the price list.
    
    Args:
        row_data: Dictionary with column values
        row_number: Row number (1-based) for error messages
    
    Returns:
        tuple: (is_valid: bool, errors: list of error messages, warnings: list of warnings)
    """
    errors = []
    warnings = []
    
    # Validate orderingNumber (required, string)
    ordering_number = row_data.get("orderingNumber")
    if ordering_number is None or str(ordering_number).strip() == "":
        errors.append(f"Row {row_number}: orderingNumber is required")
    else:
        ordering_number = str(ordering_number).strip()
        # Additional validation: ordering numbers typically have specific formats
        if len(ordering_number) < 2:
            warnings.append(f"Row {row_number}: orderingNumber '{ordering_number}' seems too short")
    
    # Validate description (required, string)
    description = row_data.get("description")
    if description is None or str(description).strip() == "":
        errors.append(f"Row {row_number}: description is required")
    
    # Validate price (required, number)
    price = row_data.get("price")
    if price is None:
        errors.append(f"Row {row_number}: price is required")
    else:
        try:
            price_value = float(price)
            if price_value < 0:
                errors.append(f"Row {row_number}: price cannot be negative (got {price_value})")
        except (ValueError, TypeError):
            errors.append(f"Row {row_number}: price must be a number (got '{price}')")
    
    # Validate SwagelokLink (optional, link)
    swagelok_link = row_data.get("SwagelokLink")
    if swagelok_link is None or str(swagelok_link).strip() == "":
        # Treat missing SwagelokLink as a warning so it is surfaced to the user,
        # but does not make the row invalid.
        warnings.append(f"Row {row_number}: SwagelokLink is missing")
    else:
        link_str = str(swagelok_link).strip()
        # Basic URL validation
        url_pattern = re.compile(
            r'^https?://'  # http:// or https://
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'  # domain
            r'localhost|'  # localhost
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # IP
            r'(?::\d+)?'  # optional port
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        if not url_pattern.match(link_str):
            warnings.append(f"Row {row_number}: SwagelokLink doesn't appear to be a valid URL: '{link_str}'")
    
    is_valid = len(errors) == 0
    return is_valid, errors, warnings


def cleanup_failed_upload(file_id, s3_key, delete_db_record=True):
    """
    Clean up S3 file and DynamoDB records when processing fails.
    
    Args:
        file_id: File ID to clean up
        s3_key: S3 key to delete
        delete_db_record: Whether to delete the file record from DynamoDB (default True)
    """
    print(f"[cleanup_failed_upload] Starting cleanup for file_id={file_id}, s3_key={s3_key}")
    
    # Delete S3 file
    try:
        s3_client.delete_object(Bucket=BUCKET, Key=s3_key)
        print(f"[cleanup_failed_upload] Deleted S3 file: {s3_key}")
    except Exception as e:
        print(f"[cleanup_failed_upload] WARNING: Failed to delete S3 file: {e}")
    
    # Delete file record from DynamoDB
    if delete_db_record:
        try:
            files_table = dynamodb.Table(FILES_TABLE)
            files_table.delete_item(Key={"fileId": file_id})
            print(f"[cleanup_failed_upload] Deleted file record: {file_id}")
        except Exception as e:
            print(f"[cleanup_failed_upload] WARNING: Failed to delete file record: {e}")
    
    # Delete all product chunks from price list products table
    try:
        products_table = dynamodb.Table(PRICE_LIST_PRODUCTS_TABLE)
        # Query all chunks for this fileId
        response = products_table.query(
            KeyConditionExpression="fileId = :fid",
            ExpressionAttributeValues={":fid": file_id},
            ProjectionExpression="fileId, chunkIndex"
        )
        items = response.get("Items", [])
        
        # Delete each chunk
        for item in items:
            products_table.delete_item(
                Key={
                    "fileId": item["fileId"],
                    "chunkIndex": item["chunkIndex"]
                }
            )
        print(f"[cleanup_failed_upload] Deleted {len(items)} product chunks for file: {file_id}")
    except Exception as e:
        print(f"[cleanup_failed_upload] WARNING: Failed to delete products: {e}")


def split_products_into_chunks(products, chunk_size=500):
    """
    Split a list of products into chunks for DynamoDB storage.
    DynamoDB has a 400KB item limit, so we chunk products to stay well under that.
    
    Args:
        products: List of product dictionaries
        chunk_size: Maximum number of products per chunk (default 500)
    
    Returns:
        list: List of product chunks (each chunk is a list of products)
    """
    if not products:
        return [[]]  # Return one empty chunk for metadata
    
    chunks = []
    for i in range(0, len(products), chunk_size):
        chunks.append(products[i:i + chunk_size])
    
    return chunks


def save_price_list_products(file_id, s3_key, products):
    """
    Save price list products to DynamoDB using chunked storage.
    Products are split into chunks to stay within DynamoDB's 400KB item limit.
    
    Table structure:
    - fileId (PK): Partition Key
    - chunkIndex (SK): Sort Key (0, 1, 2, ...)
    
    Chunk 0 contains metadata, all chunks contain products array.
    
    Args:
        file_id: File ID associated with these products
        s3_key: S3 key of the source file
        products: List of product dictionaries
    
    Returns:
        tuple: (success: bool, error_message: str or None)
    """
    table = dynamodb.Table(PRICE_LIST_PRODUCTS_TABLE)
    timestamp = int(time.time() * 1000)
    timestamp_iso = datetime.utcnow().isoformat() + 'Z'
    
    print(f"[save_price_list_products] Saving {len(products)} products for file {file_id}")
    
    try:
        # Convert float values to Decimal for DynamoDB
        products_for_db = convert_floats_to_decimal(products)

        # Log a small sample of products (without overwhelming logs) to verify SwagelokLink presence
        sample_count = min(5, len(products_for_db))
        for idx, p in enumerate(products_for_db[:sample_count]):
            try:
                print(
                    f"[save_price_list_products] Sample product {idx}: "
                    f"orderingNumber={p.get('orderingNumber')}, "
                    f"SwagelokLink={p.get('SwagelokLink')}, "
                    f"swagelokLink={p.get('swagelokLink')}"
                )
            except Exception as e:
                print(f"[save_price_list_products] WARNING: Failed to log sample product {idx}: {e}")
        
        # Split products into chunks
        chunks = split_products_into_chunks(products_for_db)
        total_chunks = len(chunks)
        
        print(f"[save_price_list_products] Splitting {len(products)} products into {total_chunks} chunks")
        
        # Save each chunk
        for chunk_idx, chunk_products in enumerate(chunks):
            item = {
                "fileId": file_id,
                "chunkIndex": chunk_idx,
                "products": chunk_products,
                "productsInChunk": len(chunk_products),
                "updatedAt": timestamp,
                "updatedAtIso": timestamp_iso,
            }
            
            # Add metadata to chunk 0
            if chunk_idx == 0:
                item["sourceFile"] = s3_key
                item["createdAt"] = timestamp
                item["createdAtIso"] = timestamp_iso
                item["totalProductsCount"] = len(products_for_db)
                item["totalChunks"] = total_chunks
            
            table.put_item(Item=item)
            print(f"[save_price_list_products] Saved chunk {chunk_idx + 1}/{total_chunks} with {len(chunk_products)} products")
        
        print(f"[save_price_list_products] SUCCESS: Saved {len(products)} products in {total_chunks} chunks")
        return True, None
        
    except Exception as e:
        error_msg = str(e)
        print(f"[save_price_list_products] ERROR: Failed to save products: {error_msg}")
        import traceback
        traceback.print_exc()
        return False, error_msg


def process_price_list(file_id, s3_key):
    """
    Process an uploaded price list (xlsx file) from S3.
    
    Workflow:
    1. Download xlsx file from S3 to temp location
    2. Validate schema (column names, order, types)
    3. Validate each row
    4. Save products to DynamoDB price list products table
    5. Update file status
    
    Args:
        file_id: File ID from S3 metadata
        s3_key: S3 key of the uploaded file
    
    Returns:
        dict: Processing result with status and product count
    """
    print(f"[process_price_list] Starting processing for file_id={file_id}, s3_key={s3_key}")
    
    try:
        # Step 1: Update status - Processing started
        update_file_status(
            file_id=file_id,
            status="processing",
            processingStage="Starting price list processing",
            s3Key=s3_key
        )
        
        # Step 2: Download xlsx file from S3 to temp location
        print(f"[process_price_list] Downloading file from S3: bucket={BUCKET}, key={s3_key}")
        
        # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
            s3_client.download_file(BUCKET, s3_key, tmp_path)
        
        print(f"[process_price_list] File downloaded to: {tmp_path}")
        
        # Step 3: Open xlsx file with openpyxl
        update_file_status(
            file_id=file_id,
            status="validating_schema",
            processingStage="Validating file schema"
        )
        
        print(f"[process_price_list] Opening xlsx file...")
        workbook = load_workbook(filename=tmp_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        if sheet is None:
            raise ValueError("No active sheet found in the xlsx file")
        
        # Step 4: Read all rows and find header row
        rows = list(sheet.iter_rows(values_only=True))
        
        if len(rows) < 2:
            raise ValueError("File must have at least a header row and one data row")
        
        # Find the header row (it might not be the first row)
        header_row_idx, headers = find_header_row(rows)
        
        if header_row_idx is None or headers is None:
            # Couldn't find header row, show what we found in first few rows
            first_rows_preview = [str(rows[i]) for i in range(min(5, len(rows)))]
            error_msg = f"Could not find header row with required columns (orderingNumber, description, price). First rows found: {first_rows_preview}"
            print(f"[process_price_list] {error_msg}")
            update_file_status(
                file_id=file_id,
                status="failed",
                processingStage="Header row not found",
                error=error_msg
            )
            os.unlink(tmp_path)
            workbook.close()
            # Cleanup failed upload
            cleanup_failed_upload(file_id, s3_key)
            return {
                "success": False,
                "fileId": file_id,
                "error": error_msg
            }
        
        print(f"[process_price_list] Found header row at index {header_row_idx}: {headers}")
        
        # Step 5: Validate schema and get column mapping
        schema_valid, schema_errors, column_mapping = validate_price_list_schema(headers)
        
        if not schema_valid:
            print(f"[process_price_list] Schema validation failed: {schema_errors}")
            error_msg = "; ".join(schema_errors)
            update_file_status(
                file_id=file_id,
                status="failed",
                processingStage="Schema validation failed",
                error=error_msg
            )
            # Clean up temp file
            os.unlink(tmp_path)
            workbook.close()
            # Cleanup failed upload
            cleanup_failed_upload(file_id, s3_key)
            return {
                "success": False,
                "fileId": file_id,
                "error": f"Schema validation failed: {error_msg}",
                "schemaErrors": schema_errors
            }
        
        print(f"[process_price_list] Schema validation passed. Column mapping: {column_mapping}")

        # Extra debug: specifically log where SwagelokLink is mapped (if present)
        if "SwagelokLink" in column_mapping:
            print(
                f"[process_price_list] SwagelokLink column mapped to index {column_mapping['SwagelokLink']}"
            )
        else:
            print(
                "[process_price_list] WARNING: SwagelokLink column NOT found in column_mapping. "
                "Links from the file will not be captured."
            )
        
        # Step 6: Process and validate each row (starting after header row)
        update_file_status(
            file_id=file_id,
            status="processing_rows",
            processingStage="Processing rows"
        )
        
        products = []
        all_errors = []
        all_warnings = []
        expected_col_names = [col["name"] for col in PRICE_LIST_SCHEMA["columns"]]
        data_start_row = header_row_idx + 1
        
        for row_idx, row in enumerate(rows[data_start_row:], start=data_start_row + 1):  # 1-based row numbers
            # Skip completely empty rows
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            
            # Create row data dictionary using column mapping
            row_data = {}
            for col_name in expected_col_names:
                if col_name in column_mapping:
                    col_idx = column_mapping[col_name]
                    if col_idx < len(row):
                        row_data[col_name] = row[col_idx]
                    else:
                        row_data[col_name] = None
                else:
                    row_data[col_name] = None

            if row_idx <= data_start_row + 5:
                # Log first few data rows to inspect SwagelokLink values
                print(
                    f"[process_price_list] Row {row_idx} mapped data: "
                    f"orderingNumber={row_data.get('orderingNumber')}, "
                    f"description={row_data.get('description')}, "
                    f"price={row_data.get('price')}, "
                    f"SwagelokLink={row_data.get('SwagelokLink')}"
                )
            
            # Validate row
            row_valid, row_errors, row_warnings = validate_price_list_row(row_data, row_idx)
            
            all_errors.extend(row_errors)
            all_warnings.extend(row_warnings)
            
            # Create product entry
            product = {
                "orderingNumber": str(row_data.get("orderingNumber", "")).strip() if row_data.get("orderingNumber") else None,
                "description": str(row_data.get("description", "")).strip() if row_data.get("description") else None,
                "price": None,
                "swagelokLink": str(row_data.get("SwagelokLink", "")).strip() if row_data.get("SwagelokLink") else None,
                "rowNumber": row_idx,
                "status": "valid" if row_valid else "invalid",
                "errors": row_errors if row_errors else None,
                "warnings": row_warnings if row_warnings else None
            }
            
            # Parse price
            try:
                if row_data.get("price") is not None:
                    product["price"] = float(row_data["price"])
            except (ValueError, TypeError):
                product["price"] = None
            
            products.append(product)

            if row_idx <= data_start_row + 5:
                # Log corresponding product representation for early rows
                print(
                    f"[process_price_list] Product built from row {row_idx}: "
                    f"orderingNumber={product.get('orderingNumber')}, "
                    f"price={product.get('price')}, "
                    f"swagelokLink={product.get('swagelokLink')}, "
                    f"status={product.get('status')}"
                )
        
        print(f"[process_price_list] Processed {len(products)} rows, {len(all_errors)} errors, {len(all_warnings)} warnings")
        
        # Clean up temp file
        os.unlink(tmp_path)
        workbook.close()
        
        # Step 7: Save products to DynamoDB
        valid_products = [p for p in products if p["status"] == "valid"]
        invalid_products = [p for p in products if p["status"] == "invalid"]
        
        update_file_status(
            file_id=file_id,
            status="saving_products",
            processingStage=f"Saving {len(products)} products to database"
        )
        
        save_success, save_error = save_price_list_products(file_id, s3_key, products)
        
        if not save_success:
            error_msg = f"Failed to save products: {save_error}"
            print(f"[process_price_list] {error_msg}")
            update_file_status(
                file_id=file_id,
                status="failed",
                processingStage="Failed to save products",
                error=error_msg
            )
            # Cleanup on failure
            cleanup_failed_upload(file_id, s3_key)
            return {
                "success": False,
                "fileId": file_id,
                "error": error_msg
            }
        
        # Step 8: Final status update
        final_status = "pending_review" if len(invalid_products) == 0 else "pending_review_with_errors"
        
        update_file_status(
            file_id=file_id,
            status=final_status,
            processingStage="Processing completed",
            productsCount=len(products),
            validProductsCount=len(valid_products),
            invalidProductsCount=len(invalid_products),
            totalErrors=len(all_errors),
            totalWarnings=len(all_warnings)
        )
        
        print(f"[process_price_list] Processing completed successfully")
        
        return {
            "success": True,
            "fileId": file_id,
            "productsCount": len(products),
            "validProductsCount": len(valid_products),
            "invalidProductsCount": len(invalid_products),
            "totalErrors": len(all_errors),
            "totalWarnings": len(all_warnings)
        }
        
    except Exception as e:
        error_message = str(e)
        print(f"[process_price_list] ERROR: {error_message}")
        import traceback
        traceback.print_exc()
        
        # Cleanup on failure
        cleanup_failed_upload(file_id, s3_key)
        
        update_file_status(
            file_id=file_id,
            status="failed",
            processingStage="Processing failed",
            error=error_message
        )
        
        return {
            "success": False,
            "fileId": file_id,
            "error": error_message
        }


def process_sales_drawing(file_id, s3_key):
    """
    Placeholder for processing sales drawing files.
    TODO: Implement when requirements are defined.
    """
    print(f"[process_sales_drawing] Processing not implemented yet for file_id={file_id}")
    
    update_file_status(
        file_id=file_id,
        status="pending_review",
        processingStage="Sales drawing processing not implemented"
    )
    
    return {
        "success": True,
        "fileId": file_id,
        "message": "Sales drawing processing not implemented"
    }


from utils.textractClient import start_job, is_job_complete, get_job_results
from utils.textractParser import (
    convert_pages_to_blocks,
    build_block_maps,
    get_special_cells_texts,
    get_header_row_count,
    convert_table_block_to_grid,
    has_ordering_number_header,
    convert_grid_to_catalog_products
)
from decimal import Decimal


def process_uploaded_file(event, context):
    """
    Process uploaded PDF file through AWS Textract to extract catalog product tables.
    Updates DynamoDB status at each step and saves products to temp table.
    
    Workflow:
    1. Parse S3 key and get file ID from S3 metadata
    2. Filter: Only process PDF files (skip JSON results)
    3. Update status: textract_started
    4. Start Textract analysis job
    5. Wait for job completion (update status periodically)
    6. Update status: textract_completed
    7. Retrieve and save Textract results to S3
    8. Update status: parsing_tables
    9. Parse table blocks from results
    10. Convert relevant tables to catalog products
    11. Save products to temp table
    12. Update status: completed with full metadata
    """
    
    # Check if this is a mid-process event (from the body)
    body = event.get('body')
    file_id = None
    s3_key = None
    start_from_mid_process = False
    textract_results_key = None
    
    if body:
        try:
            body_data = json.loads(body) if isinstance(body, str) else body
            start_from_mid_process = body_data.get('startFromMidProcess')
            if start_from_mid_process:
                s3_key = body_data.get('s3Key')
                file_id = body_data.get('fileId')
                textract_results_key = body_data.get('textractResultsKey')
                print(f"[process_uploaded_file] Mid-process event - s3Key: {s3_key}, fileId: {file_id}, textractResultsKey: {textract_results_key}")
        
        except (json.JSONDecodeError, TypeError) as e:
            print(f"[process_uploaded_file] No body, full process event: {event}")

    try:
        if start_from_mid_process and textract_results_key:
            textract_results = s3_client.get_object(Bucket=BUCKET, Key=textract_results_key)
            results_pages = json.loads(textract_results['Body'].read())
            print(f"[process_uploaded_file] Retrieved Textract results from S3: {textract_results_key}")
        
        else:
            print(f"[process_uploaded_file] Starting file processing")
            print(f"[process_uploaded_file] Event: {json.dumps(event)}")
            
            # Step 1: Parse S3 key
            s3_key = parse_s3_key(event)
            print(f"[process_uploaded_file] Parsed S3 key: {s3_key}")
            
            # Step 2: Filter - Only process PDF files, skip JSON results
            if s3_key.endswith('.json'):
                print(f"[process_uploaded_file] Skipping JSON file: {s3_key}")
                return {
                    "statusCode": 200,
                    "body": json.dumps({"message": "Skipped JSON file"}),
                }
            
            # Step 3: Get file ID from S3 object metadata
            try:
                s3_response = s3_client.head_object(Bucket=BUCKET, Key=s3_key)
                metadata = s3_response.get('Metadata', {})
                print(f"[process_uploaded_file] S3 head object response metadata: {json.dumps(metadata)}")

                file_id = metadata.get('file-id')
                file_type = metadata.get('file-type')
        
            except Exception as e:
                print(f"[process_uploaded_file] ERROR: Failed to get S3 object metadata: {e}")
                return {
                    "statusCode": 500,
                    "body": json.dumps({"error": "Failed to get S3 object metadata"}),
                }

            if not file_id or not file_type:       
                print(f"[process_uploaded_file] ERROR: No file ID or file type found")
                return {
                    "statusCode": 500,
                    "body": json.dumps({"error": "No file ID or file type found"}),
                }

            if file_type == 'Price List':
                result = process_price_list(file_id, s3_key)
                return {
                    "statusCode": 200 if result['success'] else 500,
                    "body": json.dumps(result),
                }

            elif file_type == 'Sales Drawing':
                result = process_sales_drawing(file_id, s3_key)
                return {
                    "statusCode": 200,
                    "body": json.dumps(result),
                }

            # Step 2: Update status - Textract started
            update_file_status(
                file_id=file_id,
                status="textract_started",
                processingStage="Starting Textract analysis",
                s3Key=s3_key
            )
            
            # Step 3: Start Textract job
            print(f"[process_uploaded_file] Starting Textract job for bucket={BUCKET}, key={s3_key}")
            job_id = start_job(BUCKET, s3_key, features=['TABLES'], region=AWS_REGION)
            print(f"[process_uploaded_file] Textract job started with JobId: {job_id}")
            
            update_file_status(
                file_id=file_id,
                status="textract_processing",
                processingStage="Textract analysis in progress",
                textractJobId=job_id
            )

            # Step 4: Wait for job completion with polling
            max_attempts = 60  # Increased for production (60 attempts * 2 seconds = 2 minutes max)
            attempt = 0
            status = is_job_complete(job_id, region=AWS_REGION)
            print(f"[process_uploaded_file] Initial job status: {status}")
            
            while status != "SUCCEEDED" and attempt < max_attempts:
                if status == "FAILED":
                    print(f"[process_uploaded_file] ERROR: Textract job failed after {attempt} attempts")
                    update_file_status(
                        file_id=file_id,
                        status="failed",
                        processingStage="Textract analysis failed",
                        error="Textract job failed"
                    )
                    return {
                        "statusCode": 500,
                        "body": json.dumps({"error": "Textract job failed", "fileId": file_id}),
                    }
                
                # Update status every 5 attempts (every 10 seconds)
                if attempt > 0 and attempt % 5 == 0:
                    update_file_status(
                        file_id=file_id,
                        status="textract_processing",
                        processingStage=f"Textract analysis in progress (attempt {attempt}/{max_attempts})"
                    )
                
                print(f"[process_uploaded_file] Waiting for job completion (attempt {attempt + 1}/{max_attempts})...")
                time.sleep(2)
                status = is_job_complete(job_id, region=AWS_REGION)
                print(f"[process_uploaded_file] Job status: {status}")
                attempt += 1

            if status != "SUCCEEDED":
                print(f"[process_uploaded_file] ERROR: Job timed out after {max_attempts} attempts")
                update_file_status(
                    file_id=file_id,
                    status="failed",
                    processingStage="Textract analysis timed out",
                    error=f"Job timed out after {max_attempts} attempts"
                )
                return {
                    "statusCode": 500,
                    "body": json.dumps({"error": "Job timed out", "fileId": file_id}),
                }

            print(f"[process_uploaded_file] Textract job completed successfully")
            
            # Step 5: Update status - Textract completed
            update_file_status(
                file_id=file_id,
                status="textract_completed",
                processingStage="Textract analysis completed, retrieving results"
            )
            
            # Step 6: Retrieve Textract results
            print(f"[process_uploaded_file] Retrieving Textract results...")
            results_pages = get_job_results(job_id, region=AWS_REGION)
            pages_count = len(results_pages)
            print(f"[process_uploaded_file] Retrieved {pages_count} result pages from Textract")

            # Step 7: Save Textract results to S3 as JSON
            # Format: uploads/{file_name_no_extension}_textract_results.json
            # Example: uploads/ms-02-89.pdf -> uploads/ms-02-89_textract_results.json
            base_name = s3_key.rsplit('.', 1)[0]  # Remove extension
            textract_results_key = f"{base_name}_textract_results.json"
            print(f"[process_uploaded_file] Saving Textract results to S3: bucket={BUCKET}, key={textract_results_key}")
            
            try:
                s3_client.put_object(
                    Bucket=BUCKET,
                    Key=textract_results_key,
                    Body=json.dumps(results_pages, indent=2),
                    ContentType="application/json",
                    Metadata={
                        "description": "AWS Textract analysis results",
                        "original-file": s3_key,
                        "job-id": job_id
                    }
                )
                print(f"[process_uploaded_file] Successfully saved Textract results to S3")
            except Exception as e:
                print(f"[process_uploaded_file] WARNING: Failed to save Textract results to S3: {e}")

            # Step 8: Update status - Parsing tables
            update_file_status(
                file_id=file_id,
                status="parsing_tables",
                processingStage="Parsing tables from Textract results",
                pagesCount=pages_count,
                textractResultsKey=textract_results_key
            )
        
        # Step 9: Parse blocks from Textract results
        print(f"[process_uploaded_file] Parsing blocks from Textract results...")
        all_blocks = convert_pages_to_blocks(results_pages)
        print(f"[process_uploaded_file] Total blocks parsed: {len(all_blocks)}")
        
        id_map, type_map = build_block_maps(all_blocks)
        print(f"[process_uploaded_file] Built block maps - total IDs: {len(id_map)}, block types: {list(type_map.keys())}")
        
        table_blocks = type_map.get("TABLE", [])
        tables_count = len(table_blocks)
        print(f"[process_uploaded_file] Found {tables_count} table blocks in document")

        # Step 10: Process each table block
        event_payloads = []
        next_product_id = 1  # Track product ID across all tables to ensure uniqueness
        
        for tindex, tblock in enumerate(table_blocks):
            print(f"[process_uploaded_file] Processing table {tindex + 1}/{tables_count}")
            
            # Update status for every table processed
            if tindex % 5 == 0 or tindex == 0:  # Update every 5 tables or first table
                update_file_status(
                    file_id=file_id,
                    status="parsing_tables",
                    processingStage=f"Processing table {tindex + 1} of {tables_count}"
                )
            
            # Get special cells (headers, titles, etc.)
            special_types = get_special_cells_texts(tblock, id_map)
            table_title = special_types.get('TABLE_TITLE', [{}])[0].get('text', 'N/A') if special_types.get('TABLE_TITLE') else 'N/A'
            print(f"[process_uploaded_file] Table {tindex} title: {table_title}")
            
            # Determine header row count
            header_row_index = get_header_row_count(special_types.get('COLUMN_HEADER'))
            print(f"[process_uploaded_file] Table {tindex} header row count: {header_row_index}")

            # Convert table block to grid structure
            print(f"[process_uploaded_file] Converting table {tindex} to grid structure...")
            table_grid = convert_table_block_to_grid(tblock, id_map, replicate_data=True, header_scan_rows=header_row_index)
            print(f"[process_uploaded_file] Table {tindex} grid: {len(table_grid.get('headers', []))} columns, {len(table_grid.get('rows', []))} rows")
            
            # Check if table contains ordering number column
            ordering_number_index = has_ordering_number_header(table_grid.get('headers'))
            
            if ordering_number_index is not None:
                print(f"[process_uploaded_file] Table {tindex} contains ordering number at column index {ordering_number_index}")
                print(f"[process_uploaded_file] Converting table {tindex} to catalog products (starting from ID {next_product_id})...")
                
                event_payload = convert_grid_to_catalog_products(table_grid, tblock, id_map, tindex, ordering_number_index, next_product_id)
                product_count = len(event_payload) if event_payload else 0
                print(f"[process_uploaded_file] Table {tindex} extracted {product_count} products")
                
                if event_payload:
                    event_payloads.append(event_payload)
                    # Update next_product_id to continue from where this table left off
                    next_product_id += product_count
            else:
                print(f"[process_uploaded_file] Table {tindex} does NOT contain ordering number header (Title: {table_title}) - skipping")

        tables_with_products = len(event_payloads)
        total_products = sum(len(payload) for payload in event_payloads)
        
        print(f"[process_uploaded_file] Processing complete - extracted {tables_with_products} tables with products")
        print(f"[process_uploaded_file] Total products across all tables: {total_products}")

        # Step 11: Save products to temp table
        if total_products > 0:
            update_file_status(
                file_id=file_id,
                status="saving_products",
                processingStage=f"Saving {total_products} products to database"
            )
            
            products_saved = save_products_to_catalog_products_table(file_id, s3_key, event_payloads)
            print(f"[process_uploaded_file] Saved {products_saved} products to temp table")
        else:
            products_saved = 0
            print(f"[process_uploaded_file] No products found to save")

        # Step 12: Final status update - Completed
        update_file_status(
            file_id=file_id,
            status="pending_review",
            processingStage="Processing completed successfully",
            pagesCount=pages_count,
            tablesCount=tables_count,
            tablesWithProducts=tables_with_products,
            productsCount=total_products,
            textractJobId=job_id,
            textractResultsKey=textract_results_key
        )
        
        print(f"[process_uploaded_file] File processing completed successfully for file ID: {file_id}")

        return {
            "statusCode": 200,
            "body": json.dumps({
                "fileId": file_id,
                "productsCount": total_products,
                "metadata": {
                    "pages_count": pages_count,
                    "total_tables": tables_count,
                    "tables_with_products": tables_with_products,
                    "products_count": total_products,
                    "textract_results_key": textract_results_key,
                    "job_id": job_id
                }
            }),
        }
        
    except Exception as e:
        error_message = str(e)
        print(f"[process_uploaded_file] FATAL ERROR: {error_message}")
        print(f"[process_uploaded_file] Error type: {type(e).__name__}")
        
        # Update status to failed if we have a file_id
        if file_id:
            update_file_status(
                file_id=file_id,
                status="failed",
                processingStage="Processing failed with error",
                error=error_message
            )
        
        return {
            "statusCode": 500,
            "body": json.dumps({
                "error": error_message,
                "fileId": file_id,
                "s3Key": s3_key
            }),
        }